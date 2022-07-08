import openpyxl
# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("test.xlsx")
# Define variable to read the active sheet:
worksheet = wookbook.active

# Iterate the loop to read the cell values
for i in range(0, worksheet.max_row):
    mas = []
    for col in worksheet.iter_cols(1, worksheet.max_column):
        # print(col[i].value, end="\t\t")
        mas.append(col[i].value)
    print(f'\telif income == \'{mas[1]}\':')
    print(f'\t\tnv.say(\'{mas[0]}\')')



# elif work_address == 'Новосибирская обл, Коченевский р-н, рп Чик':
#         nv.say('address_chik')